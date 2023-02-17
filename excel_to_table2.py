from openpyxl import Workbook, load_workbook
import psycopg2
import sys

# 0. Define all the variable whose values can be modified
'''
    Change the value of excelFile to be the location of your excel file
    Change the value of sheetName to be name of the sheet
    Change the the value of hostname, database, username, pwd, and port_id if needed
    Change the value of tableName to be your desired name for the table
'''
excelFile = 'demoexcel.xlsx'
sheetName = 'Sheet1'
hostname = 'localhost'
database = 'project'
username = 'postgres'
pwd = 'password'
port_id = 5432
tableName = 'workdayTable'

# sh is the variable that points to the excel sheet
sh = load_workbook(excelFile)[sheetName]
conn = None

# get number of rows in the excel sheet
row_ct = 0
for max_row, row in enumerate(sh, 1):
    if not all(col.value is None for col in row):
        row_ct += 1

# get number of columns in the excel sheet
col_ct = 0
for i in range(1, sh.max_column + 1):
    if sh.cell(1, i).value == None:
        break
    col_ct = i

# check dimensions
print('Check if the dimensions are correct')
print('Row count: ' + str(row_ct) + '\tColumn count: ' + str(col_ct))
print('\n')

# return if the first column is empty
if col_ct == 0:
    print('The first column is empty.')
    sys.exit()

# 1. connect to the database
try:
        with psycopg2.connect(
                    host = hostname,
                    dbname = database,
                    user = username,
                    password = pwd,
                    port = port_id) as conn:

            # create a cursor to execute SQL code
            with conn.cursor() as cur:

                # drop the the table if it already exists
                cur.execute('DROP TABLE IF EXISTS ' + tableName)
                
# 2. Get the values of the column names
                col_names = []
                for column in range(1, col_ct + 1):
                    col_names.append(str(sh.cell(1, column).internal_value))

                # replace illegal characters with underscores
                for i in range(len(col_names)):
                    for j in range(len(col_names[i])):
                        if not col_names[i][j].isalpha() and not col_names[i][j].isnumeric() and not col_names[i][j] == '_':
                            col_names[i] = col_names[i][:j] + '_' + col_names[i][j+1:]

                # fix column names that do not start with letters
                for i in range(len(col_names)):
                    if not col_names[i][0].isalpha():
                        col_names[i] = 'x_' + col_names[i]

                # check col_names
                print('Check if the list of column names is correct')
                print(col_names)
                print('\n')

# 3. Get type of each column
                col_types = []
                for column in range(1, col_ct + 1):
                    isBool = False
                    isNum = False
                    for row in range(2, row_ct + 1):
                        # check if type in the cell is not NoneType
                        if sh.cell(row, column).internal_value:
                            # check if type in the cell is a boolean
                            if str(sh.cell(row, column).internal_value).lower() == 'true' or str(sh.cell(row, column).internal_value).lower() == 'false':
                                isBool = True
                            # check if type in the cell is a numeric
                            elif type(sh.cell(row, column).internal_value) == int or type(sh.cell(row, column).internal_value) == float:
                                isNum = True
                            # conditions for when the column is a varchar
                            if isBool and isNum or type(sh.cell(row, column).internal_value) == str:
                                col_types.append('varchar')
                                break
                        if row == row_ct and not isBool and not isNum:
                            col_types.append('varchar')
                        elif row == row_ct and isBool:
                            col_types.append('boolean')
                        elif row == row_ct and isNum:
                            col_types.append('numeric')
                        
                # check col_type
                print('Check if the list of column types is correct')
                print(col_types)
                print('\n')

                # Get the size of each column
                col_lengths = []
                for column in range(1, col_ct + 1):
                    # If type is string, get the length of the longest string in the column
                    if col_types[column - 1] == 'varchar':
                        max_length = 1
                        for row in range(2, row_ct + 1):
                            if len(str(sh.cell(row, column).internal_value)) > max_length:
                               max_length = len(str(sh.cell(row, column).internal_value))
                        col_lengths.append(max_length)
                    # If type is a boolean, do nothing
                    else:
                        col_lengths.append("skip")

                # check col_length
                print('Check if the list of length of the longest strings is correct')
                print(col_lengths)
                print('\n')
                    
# 4. Create the table structure
                create_table_script = 'Create Table ' + tableName + ' ('
                for i in range(len(col_names)):
                    create_table_script += col_names[i]
                    if col_types[i] == 'varchar':
                        create_table_script += ' varchar(' + str(col_lengths[i]) + '), '
                    elif col_types[i] == 'boolean':
                        create_table_script += ' boolean, '
                    else:
                        create_table_script += ' numeric, '
                create_table_script = create_table_script[:-2] + ")"
                cur.execute(create_table_script)

                # check create_table_script
                print('Check if the SQL code for creating the table is correct')
                print(create_table_script)
                print('\n')

# 5. Insert values into the table
                print('Check if the SQL code for inserting to the table is correct')
                for row in range(2, row_ct + 1):
                    insert_script = 'INSERT INTO ' + tableName + ' ('
                    for i in range(len(col_names)):
                        insert_script += col_names[i] + ', '
                    insert_script = insert_script[:-2] + ') VALUES ('
                    for column in range(1, col_ct + 1):
                        if str(sh.cell(row, column).internal_value).lower() != 'false' and not sh.cell(row, column).internal_value:
                            insert_script += 'NULL, '
                        else:
                            insert_script += '\'' + str(sh.cell(row, column).internal_value).replace("'", "''") + '\', '
                    
                            
                    insert_script = insert_script[:-2] + ')'
                    cur.execute(insert_script)
                    
                    # check insert_script
                    print(insert_script)

except Exception as error:
    print(error)
finally:
    if conn is not None:
        conn.close()

